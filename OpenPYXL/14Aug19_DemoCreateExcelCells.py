from openpyxl import Workbook

book = Workbook()
sheet = book.active

sheet['b4'] = 'Python'
sheet['c5'] = 'Sheet'
sheet.cell(row=5, column=5).value = "Python cells"
sheet.cell(row=6, column=6).value = "Python cells 6"

# for i in range(5):
#     for j in range(5):
#         sheet.cell(row=i, column=j).value = i + j

book.save('ExcelCells.xlsx')