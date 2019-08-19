import openpyxl

book = openpyxl.load_workbook('iterbycols.xlsx')
sheet = book.active

book.create_sheet('May')
print("Sheet Name : ", book.sheetnames)

janSheet = book['January']
book.remove(janSheet)

print("Sheet Name : ", book.sheetnames)

book.create_sheet('January', 1)
print("Sheet Name : ", book.sheetnames)

sheet = book["March"]
sheet.sheet_properties.tabColor = "0072BA"

book.save('AddNewSheet.xlsx')
